'https://www.sigma-explorer.com/explorer/map/data_map.php?year=2020&indis=pw&modi=pw_t&regi=WOR'
#data are from

# path to save
import os
os.path.abspath("test")
print(os.getcwd())
os.chdir('C:\\Users\\Riho\\Desktop\\essay\\try')

import requests
import re
import openpyxl     #Excel module

def q1(url):       #get head of my computer
    head = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:93.0) Gecko/20100101 Firefox/93.0',
        'Accept-Language': 'zh-CN,zh;q=0.8,zh-TW;q=0.7,zh-HK;q=0.5,en-US;q=0.3,en;q=0.2',
        # 'Referer':'https://stock.finance.sina.com.cn/'
        }
    proxies = {'http': 'http://localhost:7890', 'https': 'http://localhost:7890'}
    req=requests.get(url,headers=head)
    # req.encoding=req.apparent_encoding
    return req.text

def q3(life_if,yaer):
    wb = openpyxl.Workbook()  # instance
    q100 = wb.active  # create table
    q100.cell(column=1, row=1, value=str('name'))  # title
    q100.cell(column=2, row=1, value=str('value'))  # value

    if life_if == 'Life':
        url=f'https://www.sigma-explorer.com/explorer/map/data_map.php?year={yaer}&indis=pw&modi=pw_l&regi=WOR'        #Life
    elif life_if == 'Non-Life':
        url=f'https://www.sigma-explorer.com/explorer/map/data_map.php?year={yaer}&indis=pw&modi=pw_nl&regi=WOR'  # non_life
    elif life_if == 'GDP'   :
        url=f'https://www.sigma-explorer.com/explorer/map/data_map.php?year={yaer}&indis=ynd&modi=ynd&regi=WOR'  #GDP in USD bn
    elif life_if == 'GDP_per_capita' :
        url=f'https://www.sigma-explorer.com/explorer/map/data_map.php?year={yaer}&indis=yrbd&modi=yrbd&regi=WOR' #GDP per capita
    elif life_if == 'Premium_growth' :
        url=f'https://www.sigma-explorer.com/explorer/map/data_map.php?year={yaer}&indis=rpgr&modi=tot_rpgr&regi=WOR' #Premium growth
    elif life_if == 'LifePremium_density' :
        url=f'https://www.sigma-explorer.com/explorer/map/data_map.php?year={yaer}&indis=dens&modi=life_dens&regi=WOR' #Life Premium Density    
    elif life_if == 'NonLifePremium_density' :
        url=f'https://www.sigma-explorer.com/explorer/map/data_map.php?year={yaer}&indis=dens&modi=nlife_dens&regi=WOR' #NonLife Premium Density
    elif life_if == 'LifePremium_penetration' :
        url=f'https://www.sigma-explorer.com/explorer/map/data_map.php?year={yaer}&indis=pene&modi=life_pene&regi=WOR' #Life Premium Penetration
    elif life_if == 'NonLifePremium_penetration' :
        url=f'https://www.sigma-explorer.com/explorer/map/data_map.php?year={yaer}&indis=pene&modi=nlife_pene&regi=WOR' #NonLife Premium Penetration    
    else:
        return 'error'
    html=q1(url)
    print(html)
    html=html.split('},{')
    # using Re-expression to clean json dataset
    for i in html:

        key=re.findall("'hc-key': '(.*?)'",str(i))
        print('name',key[0])

        if key[0] == '':
            continue

        val =re.findall("'value': (\d+\.\d+)?",str(i))
        print('value',val[0])

        q100.append([key[0],val[0]])

    wb.save(f'{life_if}{yaer}.xlsx')


# get premium data
if __name__ == '__main__':
    for papg in range(2011,2021):
        print(papg)
        q2('Non-Life',papg)
        
#get GDP data
if __name__ == '__main__':
    for papg in range(2011,2021):
        print(papg)
        q3('GDP',papg)

# GDP per capita
if __name__ == '__main__':
    for papg in range(2011,2021):
        print(papg)
        q3('GDP_per_capita',papg)
        
#premium growth
if __name__ == '__main__':
    for papg in range(2011,2021):
        print(papg)
        q3('Premium_growth',papg)        
        
#LifePremium_density
if __name__ == '__main__':
    for papg in range(2011,2021):
        print(papg)
        q3('LifePremium_density',papg)         
        
#NonLifePremium_density
if __name__ == '__main__':
    for papg in range(2011,2021):
        print(papg)
        q3('NonLifePremium_density',papg)      
        
#LifePremium_penetration
if __name__ == '__main__':
    for papg in range(2011,2021):
        print(papg)
        q3('LifePremium_penetration',papg)          
        
#NonLifePremium_penetration
if __name__ == '__main__':
    for papg in range(2011,2021):
        print(papg)
        q3('NonLifePremium_penetration',papg)             
        
        
        